"""
SecQ-Auto CowAgent — Main FastAPI Application
Autonomous Vendor Security Questionnaire Automation
"""

import os
import time
import uuid
import hashlib
import hmac
import json
import logging
from datetime import datetime, timedelta
from typing import Optional
from contextlib import asynccontextmanager

import httpx
from fastapi import FastAPI, HTTPException, Depends, Header, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

# ============================================
# Configuration
# ============================================

NORTHFLANK_URL = os.getenv("NORTHFLANK_URL", "https://your-app.northflank.app")
OMNI_URL = os.getenv("OMNI_URL", f"{NORTHFLANK_URL}:3000/v1/chat/completions")
ROUTER9_URL = os.getenv("ROUTER9_URL", f"{NORTHFLANK_URL}:4000/v1/chat/completions")
QDRANT_URL = os.getenv("QDRANT_URL", f"{NORTHFLANK_URL}:6333")
JWT_SECRET = os.getenv("JWT_SECRET", "change-this-in-production")
EMBEDDING_URL = os.getenv("EMBEDDING_URL", f"{NORTHFLANK_URL}:3000/v1/embeddings")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("cowagent")

# ============================================
# In-Memory Stores (replace with real DB in production)
# ============================================

tenants: dict[str, dict] = {}
questionnaires: dict[str, dict] = {}
questions_store: dict[str, list] = {}
answers_store: dict[str, dict] = {}
kb_docs: dict[str, dict] = {}
kb_chunks: dict[str, list] = {}
api_keys: dict[str, str] = {}  # api_key -> tenant_id

# ============================================
# Pydantic Models
# ============================================

class TenantCreate(BaseModel):
    name: str
    slug: str
    plan: str = "trial"
    company_name: str = ""
    company_domain: str = ""
    compliance_frameworks: list[str] = []

class TenantResponse(BaseModel):
    id: str
    name: str
    slug: str
    plan: str
    api_key: str
    created_at: str

class QuestionnaireCreate(BaseModel):
    name: str
    description: str = ""
    source_format: str = "excel"

class QuestionnaireResponse(BaseModel):
    id: str
    tenant_id: str
    name: str
    status: str
    progress: dict
    created_at: str

class QuestionResponse(BaseModel):
    id: str
    questionnaire_id: str
    section: str
    number: str
    text: str
    type: str
    required: bool
    options: list[str] = []
    answer: Optional[dict] = None

class AnswerRequest(BaseModel):
    question_ids: list[str] = []  # empty = all unanswered

class AnswerResponse(BaseModel):
    id: str
    question_id: str
    text: str
    confidence: float
    citations: list[dict]
    reasoning: str
    status: str
    model_used: str
    tokens_used: int
    processing_time_ms: int

class ValidateRequest(BaseModel):
    answer_ids: list[str] = []

class ValidateResponse(BaseModel):
    answer_id: str
    status: str  # pass / flag / fail
    issues: list[str]
    suggestions: list[str]

class KBIngestRequest(BaseModel):
    name: str
    type: str = "policy"
    description: str = ""
    tags: list[str] = []
    content: str = ""  # raw text content

class ChatRequest(BaseModel):
    message: str
    questionnaire_id: str = ""

class HealthResponse(BaseModel):
    status: str
    version: str
    uptime_seconds: float
    services: dict

# ============================================
# Authentication
# ============================================

def create_api_key() -> str:
    return f"sqa_{uuid.uuid4().hex[:32]}"

def hash_api_key(key: str) -> str:
    return hashlib.sha256(key.encode()).hexdigest()

async def verify_tenant(x_api_key: str = Header(...)) -> str:
    """Verify API key and return tenant_id."""
    if x_api_key in api_keys:
        return api_keys[x_api_key]
    raise HTTPException(status_code=401, detail="Invalid API key")

# ============================================
# Agent Classes
# ============================================

class ParserAgent:
    """Parses questionnaire files into structured questions."""

    @staticmethod
    async def parse_excel(content: bytes, questionnaire_id: str) -> list[dict]:
        """Parse Excel questionnaire."""
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content))
            ws = wb.active
            questions = []
            current_section = ""
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # skip header
                if not any(cell for cell in row if cell):
                    continue
                # Detect section headers (single cell rows)
                if row[0] and not any(row[1:]):
                    current_section = str(row[0])
                    continue
                q = {
                    "id": f"q_{uuid.uuid4().hex[:12]}",
                    "questionnaire_id": questionnaire_id,
                    "section": current_section,
                    "number": str(row[0]) if row[0] else str(i),
                    "text": str(row[1]) if row[1] else "",
                    "type": ParserAgent._detect_type(row[2] if len(row) > 2 else None),
                    "required": str(row[3]).lower() in ("yes", "true", "1", "required") if len(row) > 3 else True,
                    "options": [o.strip() for o in str(row[4]).split(";")] if len(row) > 4 and row[4] else [],
                    "context": str(row[5]) if len(row) > 5 and row[5] else "",
                }
                if q["text"]:
                    questions.append(q)
            return questions
        except ImportError:
            logger.warning("openpyxl not installed, returning sample questions")
            return ParserAgent._sample_questions(questionnaire_id)

    @staticmethod
    async def parse_csv(content: bytes, questionnaire_id: str) -> list[dict]:
        """Parse CSV questionnaire."""
        import csv
        from io import StringIO
        text = content.decode("utf-8")
        reader = csv.DictReader(StringIO(text))
        questions = []
        current_section = ""
        for i, row in enumerate(reader):
            if "Section" in row and row["Section"] and "Question" not in row:
                current_section = row["Section"]
                continue
            q_text = row.get("Question", row.get("question", ""))
            if not q_text:
                continue
            q = {
                "id": f"q_{uuid.uuid4().hex[:12]}",
                "questionnaire_id": questionnaire_id,
                "section": row.get("Section", row.get("section", current_section)),
                "number": row.get("Number", row.get("number", str(i + 1))),
                "text": q_text,
                "type": ParserAgent._detect_type(row.get("Type", row.get("type", ""))),
                "required": str(row.get("Required", "yes")).lower() in ("yes", "true", "1"),
                "options": [o.strip() for o in row.get("Options", "").split(";") if o.strip()],
                "context": row.get("Context", row.get("context", "")),
            }
            questions.append(q)
        return questions

    @staticmethod
    async def parse_pdf(content: bytes, questionnaire_id: str) -> list[dict]:
        """Parse PDF questionnaire."""
        try:
            import pdfplumber
            from io import BytesIO
            questions = []
            with pdfplumber.open(BytesIO(content)) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    for i, line in enumerate(text.split("\n")):
                        line = line.strip()
                        if line and len(line) > 10:
                            q = {
                                "id": f"q_{uuid.uuid4().hex[:12]}",
                                "questionnaire_id": questionnaire_id,
                                "section": "",
                                "number": str(i + 1),
                                "text": line,
                                "type": "free_text",
                                "required": True,
                                "options": [],
                                "context": "",
                            }
                            questions.append(q)
            return questions
        except ImportError:
            logger.warning("pdfplumber not installed, returning sample questions")
            return ParserAgent._sample_questions(questionnaire_id)

    @staticmethod
    def _detect_type(type_str) -> str:
        if not type_str:
            return "free_text"
        t = str(type_str).lower()
        if "yes" in t or "bool" in t:
            return "yes_no"
        elif "choice" in t or "select" in t or "multi" in t:
            return "multiple_choice"
        elif "number" in t or "numeric" in t:
            return "numeric"
        elif "date" in t:
            return "date"
        elif "evidence" in t or "upload" in t:
            return "evidence_request"
        return "free_text"

    @staticmethod
    def _sample_questions(questionnaire_id: str) -> list[dict]:
        return [
            {
                "id": f"q_{uuid.uuid4().hex[:12]}",
                "questionnaire_id": questionnaire_id,
                "section": "Access Control",
                "number": "1.1",
                "text": "Does your organization implement multi-factor authentication for all administrative access?",
                "type": "yes_no",
                "required": True,
                "options": ["Yes", "No"],
                "context": "Access control and authentication",
            },
            {
                "id": f"q_{uuid.uuid4().hex[:12]}",
                "questionnaire_id": questionnaire_id,
                "section": "Access Control",
                "number": "1.2",
                "text": "Describe your password policy including minimum length, complexity, and rotation requirements.",
                "type": "free_text",
                "required": True,
                "options": [],
                "context": "Password management",
            },
            {
                "id": f"q_{uuid.uuid4().hex[:12]}",
                "questionnaire_id": questionnaire_id,
                "section": "Data Protection",
                "number": "2.1",
                "text": "What encryption standard is used for data at rest?",
                "type": "multiple_choice",
                "required": True,
                "options": ["AES-128", "AES-256", "RSA-2048", "RSA-4096", "Other"],
                "context": "Encryption standards",
            },
            {
                "id": f"q_{uuid.uuid4().hex[:12]}",
                "questionnaire_id": questionnaire_id,
                "section": "Data Protection",
                "number": "2.2",
                "text": "Is data encrypted in transit using TLS 1.2 or higher?",
                "type": "yes_no",
                "required": True,
                "options": ["Yes", "No"],
                "context": "Data in transit",
            },
            {
                "id": f"q_{uuid.uuid4().hex[:12]}",
                "questionnaire_id": questionnaire_id,
                "section": "Compliance",
                "number": "3.1",
                "text": "Does your organization hold any compliance certifications? Please list all that apply.",
                "type": "free_text",
                "required": True,
                "options": [],
                "context": "Compliance and certifications",
            },
        ]


class RetrievalAgent:
    """Retrieves relevant context from knowledge base using Qdrant."""

    @staticmethod
    async def search(query: str, tenant_id: str, top_k: int = 5) -> list[dict]:
        """Search Qdrant for relevant chunks."""
        try:
            # Get embedding for query
            embedding = await RetrievalAgent._get_embedding(query)
            if not embedding:
                return []

            # Search Qdrant
            async with httpx.AsyncClient(timeout=10.0) as client:
                resp = await client.post(
                    f"{QDRANT_URL}/collections/kb-chunks/points/search",
                    json={
                        "vector": embedding,
                        "filter": {"must": [{"key": "tenant_id", "match": {"value": tenant_id}}]},
                        "limit": top_k,
                        "with_payload": True,
                        "with_vector": False,
                    },
                )
                if resp.status_code == 200:
                    results = resp.json().get("result", [])
                    return [
                        {
                            "id": r["id"],
                            "text": r["payload"].get("text", ""),
                            "score": r["score"],
                            "doc_id": r["payload"].get("doc_id", ""),
                            "doc_name": r["payload"].get("doc_name", ""),
                            "chunk_index": r["payload"].get("chunk_index", 0),
                        }
                        for r in results
                    ]
                return []
        except Exception as e:
            logger.error(f"Retrieval error: {e}")
            return []

    @staticmethod
    async def _get_embedding(text: str) -> list[float]:
        """Get embedding vector from embedding service."""
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.post(
                    EMBEDDING_URL,
                    json={"model": "text-embedding-3-small", "input": text[:8000]},
                )
                if resp.status_code == 200:
                    data = resp.json()
                    return data["data"][0]["embedding"]
        except Exception as e:
            logger.error(f"Embedding error: {e}")
        return []


class AnswerAgent:
    """Generates answers using LLM with RAG context."""

    SYSTEM_PROMPT = """You are a senior security compliance analyst answering vendor security questionnaires. Your responses must be accurate, evidence-based, and professionally written.

Rules:
1. Answer the question directly and precisely
2. Only use information from the provided context
3. If confidence is below 0.7, say "Insufficient information to answer confidently"
4. Match the question type exactly (yes/no for yes_no questions, select from options for multiple_choice)
5. Cite specific evidence for every claim
6. Use the company's exact terminology
7. Never hallucinate or make up information
8. Flag when evidence is outdated or conflicting

Format your response as JSON:
{
  "answer": "your precise answer",
  "confidence": 0.0-1.0,
  "citations": [{"doc_name": "...", "excerpt": "...", "relevance": 0.0-1.0}],
  "reasoning": "how you derived this answer"
}"""

    @staticmethod
    async def generate_answer(
        question: dict,
        context_chunks: list[dict],
        tenant_profile: dict = None,
    ) -> dict:
        """Generate an answer for a question using retrieved context."""
        start_time = time.time()

        # Build context string
        context_str = "\n\n".join(
            f"[Source: {c.get('doc_name', 'Unknown')} (relevance: {c['score']:.2f})]\n{c['text']}"
            for c in context_chunks
        ) if context_chunks else "No relevant context found in knowledge base."

        user_message = f"""Question: {question['text']}
Question Type: {question['type']}
Section: {question['section']}
Required: {'Yes' if question['required'] else 'No'}
{f"Options: {', '.join(question['options'])}" if question.get('options') else ""}

Relevant Context from Knowledge Base:
{context_str}

Please provide a precise, evidence-based answer."""

        # Try OmniRoute first, fall back to 9Router
        result = await AnswerAgent._call_llm(
            AnswerAgent.SYSTEM_PROMPT, user_message, OMNI_URL
        )
        if not result:
            logger.warning("OmniRoute failed, falling back to 9Router")
            result = await AnswerAgent._call_llm(
                AnswerAgent.SYSTEM_PROMPT, user_message, ROUTER9_URL
            )

        elapsed_ms = int((time.time() - start_time) * 1000)

        if result:
            try:
                parsed = json.loads(result)
                return {
                    "text": parsed.get("answer", result),
                    "confidence": parsed.get("confidence", 0.5),
                    "citations": parsed.get("citations", []),
                    "reasoning": parsed.get("reasoning", ""),
                    "model_used": "omniroute/9router",
                    "tokens_used": 0,
                    "processing_time_ms": elapsed_ms,
                }
            except json.JSONDecodeError:
                return {
                    "text": result[:2000],
                    "confidence": 0.5,
                    "citations": [],
                    "reasoning": "Raw LLM response (JSON parse failed)",
                    "model_used": "omniroute/9router",
                    "tokens_used": 0,
                    "processing_time_ms": elapsed_ms,
                }

        return {
            "text": "Unable to generate answer. Please try again or answer manually.",
            "confidence": 0.0,
            "citations": [],
            "reasoning": "LLM call failed on all providers",
            "model_used": "none",
            "tokens_used": 0,
            "processing_time_ms": elapsed_ms,
        }

    @staticmethod
    async def _call_llm(system: str, user: str, url: str) -> Optional[str]:
        """Call LLM API with fallback."""
        try:
            async with httpx.AsyncClient(timeout=60.0) as client:
                resp = await client.post(
                    url,
                    json={
                        "model": "auto",
                        "messages": [
                            {"role": "system", "content": system},
                            {"role": "user", "content": user},
                        ],
                        "max_tokens": 1024,
                        "temperature": 0.1,
                    },
                )
                if resp.status_code == 200:
                    data = resp.json()
                    return data["choices"][0]["message"]["content"]
                logger.warning(f"LLM call failed: {resp.status_code} from {url}")
        except Exception as e:
            logger.error(f"LLM call error: {e} from {url}")
        return None


class ValidatorAgent:
    """Validates generated answers for quality and accuracy."""

    VALIDATION_PROMPT = """You are a quality assurance reviewer for security questionnaire answers. Validate the following answer against the question and context.

Check:
1. Does the answer directly address the question?
2. Is the confidence score calibrated correctly?
3. Are citations relevant and sufficient?
4. Are there contradictions in the evidence?
5. Does the answer use appropriate professional language?
6. Are there compliance risks (over-promising, inaccurate claims)?

Return JSON:
{
  "status": "pass" | "flag" | "fail",
  "issues": ["list of specific issues"],
  "suggestions": ["list of suggested improvements"]
}"""

    @staticmethod
    async def validate_answer(question: dict, answer: dict, context_chunks: list[dict]) -> dict:
        """Validate a generated answer."""
        user_message = f"""Question: {question['text']}
Question Type: {question['type']}
Required: {'Yes' if question['required'] else 'No'}

Proposed Answer: {answer.get('text', '')}
Confidence: {answer.get('confidence', 0)}
Reasoning: {answer.get('reasoning', '')}

Context Used:
{chr(10).join(c.get('text', '')[:200] for c in context_chunks)}"""

        try:
            result = await AnswerAgent._call_llm(
                ValidatorAgent.VALIDATION_PROMPT, user_message, OMNI_URL
            )
            if result:
                try:
                    parsed = json.loads(result)
                    return {
                        "status": parsed.get("status", "flag"),
                        "issues": parsed.get("issues", []),
                        "suggestions": parsed.get("suggestions", []),
                    }
                except json.JSONDecodeError:
                    pass
        except Exception as e:
            logger.error(f"Validation error: {e}")

        # Default validation
        confidence = answer.get("confidence", 0)
        if confidence >= 0.8:
            return {"status": "pass", "issues": [], "suggestions": []}
        elif confidence >= 0.5:
            return {"status": "flag", "issues": ["Low confidence"], "suggestions": ["Review and verify"]}
        return {"status": "fail", "issues": ["Very low confidence"], "suggestions": ["Manual review required"]}


class FormatterAgent:
    """Formats completed questionnaires for export."""

    @staticmethod
    def format_excel(questionnaire: dict, questions: list[dict], answers: dict[str, dict]) -> bytes:
        """Export questionnaire as Excel."""
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Questionnaire"
            headers = ["Section", "Number", "Question", "Type", "Required", "Answer", "Confidence", "Status"]
            ws.append(headers)
            for q in questions:
                a = answers.get(q["id"], {})
                ws.append([
                    q.get("section", ""),
                    q.get("number", ""),
                    q.get("text", ""),
                    q.get("type", ""),
                    "Yes" if q.get("required") else "No",
                    a.get("text", ""),
                    a.get("confidence", ""),
                    a.get("status", "unanswered"),
                ])
            buf = BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except ImportError:
            # Fallback to CSV
            return FormatterAgent.format_csv(questionnaire, questions, answers)

    @staticmethod
    def format_csv(questionnaire: dict, questions: list[dict], answers: dict[str, dict]) -> bytes:
        """Export questionnaire as CSV."""
        import csv
        from io import StringIO
        buf = StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Section", "Number", "Question", "Type", "Required", "Answer", "Confidence", "Status"])
        for q in questions:
            a = answers.get(q["id"], {})
            writer.writerow([
                q.get("section", ""),
                q.get("number", ""),
                q.get("text", ""),
                q.get("type", ""),
                "Yes" if q.get("required") else "No",
                a.get("text", ""),
                a.get("confidence", ""),
                a.get("status", "unanswered"),
            ])
        return buf.getvalue().encode("utf-8")


# ============================================
# Application Lifecycle
# ============================================

start_time = time.time()

@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("🐄 CowAgent starting up...")
    # Create default tenant for testing
    tenant_id = "tenant_default"
    default_key = "sqa_default_key_change_me"
    tenants[tenant_id] = {
        "id": tenant_id,
        "name": "Default Tenant",
        "slug": "default",
        "plan": "trial",
        "company_name": "Acme Corp",
        "company_domain": "acme.com",
        "compliance_frameworks": ["SOC2", "ISO27001"],
        "created_at": datetime.utcnow().isoformat(),
    }
    api_keys[default_key] = tenant_id
    logger.info(f"Default tenant created: {tenant_id}")
    logger.info(f"Default API key: {default_key}")
    yield
    logger.info("🐄 CowAgent shutting down...")


# ============================================
# FastAPI Application
# ============================================

app = FastAPI(
    title="SecQ-Auto CowAgent",
    description="Autonomous Vendor Security Questionnaire Automation",
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================
# Health & Status
# ============================================

@app.get("/health", response_model=HealthResponse)
async def health_check():
    return HealthResponse(
        status="healthy",
        version="1.0.0",
        uptime_seconds=time.time() - start_time,
        services={
            "omniroute": OMNI_URL,
            "router9": ROUTER9_URL,
            "qdrant": QDRANT_URL,
        },
    )


@app.get("/")
async def root():
    return {
        "name": "SecQ-Auto CowAgent",
        "version": "1.0.0",
        "docs": "/docs",
        "health": "/health",
    }


# ============================================
# Tenant Management
# ============================================

@app.post("/api/tenants", response_model=TenantResponse)
async def create_tenant(tenant: TenantCreate):
    """Create a new tenant (multi-tenancy)."""
    tenant_id = f"tenant_{uuid.uuid4().hex[:12]}"
    api_key = create_api_key()
    now = datetime.utcnow().isoformat()
    tenants[tenant_id] = {
        "id": tenant_id,
        "name": tenant.name,
        "slug": tenant.slug,
        "plan": tenant.plan,
        "company_name": tenant.company_name,
        "company_domain": tenant.company_domain,
        "compliance_frameworks": tenant.compliance_frameworks,
        "created_at": now,
    }
    api_keys[api_key] = tenant_id
    logger.info(f"Tenant created: {tenant_id} ({tenant.slug})")
    return TenantResponse(
        id=tenant_id,
        name=tenant.name,
        slug=tenant.slug,
        plan=tenant.plan,
        api_key=api_key,
        created_at=now,
    )


# ============================================
# Questionnaire Management
# ============================================

@app.post("/api/questionnaires", response_model=QuestionnaireResponse)
async def create_questionnaire(
    q: QuestionnaireCreate,
    tenant_id: str = Depends(verify_tenant),
):
    """Create a new questionnaire."""
    q_id = f"q_{uuid.uuid4().hex[:12]}"
    now = datetime.utcnow().isoformat()
    questionnaires[q_id] = {
        "id": q_id,
        "tenant_id": tenant_id,
        "name": q.name,
        "description": q.description,
        "source_format": q.source_format,
        "status": "uploading",
        "progress": {"total": 0, "parsed": 0, "answered": 0, "validated": 0, "approved": 0},
        "created_at": now,
    }
    questions_store[q_id] = []
    answers_store[q_id] = {}
    return QuestionnaireResponse(
        id=q_id,
        tenant_id=tenant_id,
        name=q.name,
        status="uploading",
        progress=questionnaires[q_id]["progress"],
        created_at=now,
    )


@app.post("/api/questionnaires/{q_id}/upload")
async def upload_questionnaire(
    q_id: str,
    file: UploadFile = File(...),
    tenant_id: str = Depends(verify_tenant),
):
    """Upload and parse a questionnaire file."""
    if q_id not in questionnaires:
        raise HTTPException(status_code=404, detail="Questionnaire not found")
    if questionnaires[q_id]["tenant_id"] != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")

    content = await file.read()
    filename = file.filename or "unknown"

    # Parse based on file type
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        questions = await ParserAgent.parse_excel(content, q_id)
    elif filename.endswith(".csv"):
        questions = await ParserAgent.parse_csv(content, q_id)
    elif filename.endswith(".pdf"):
        questions = await ParserAgent.parse_pdf(content, q_id)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {filename}")

    questions_store[q_id] = questions
    questionnaires[q_id]["status"] = "ready"
    questionnaires[q_id]["progress"]["total"] = len(questions)
    questionnaires[q_id]["progress"]["parsed"] = len(questions)

    return {
        "questionnaire_id": q_id,
        "questions_parsed": len(questions),
        "status": "ready",
    }


@app.get("/api/questionnaires/{q_id}/questions")
async def get_questions(
    q_id: str,
    tenant_id: str = Depends(verify_tenant),
):
    """Get all questions for a questionnaire."""
    if q_id not in questionnaires:
        raise HTTPException(status_code=404, detail="Questionnaire not found")
    if questionnaires[q_id]["tenant_id"] != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")

    qs = questions_store.get(q_id, [])
    result = []
    for q in qs:
        a = answers_store.get(q_id, {}).get(q["id"])
        result.append({**q, "answer": a})
    return {"questions": result, "total": len(result)}


@app.post("/api/questionnaires/{q_id}/answer")
async def generate_answers(
    q_id: str,
    req: AnswerRequest,
    tenant_id: str = Depends(verify_tenant),
):
    """Generate AI answers for questions."""
    if q_id not in questionnaires:
        raise HTTPException(status_code=404, detail="Questionnaire not found")
    if questionnaires[q_id]["tenant_id"] != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")

    qs = questions_store.get(q_id, [])
    if req.question_ids:
        qs = [q for q in qs if q["id"] in req.question_ids]

    tenant = tenants.get(tenant_id, {})
    generated = []

    for q in qs:
        # Retrieve context
        chunks = await RetrievalAgent.search(q["text"], tenant_id, top_k=5)

        # Generate answer
        answer_data = await AnswerAgent.generate_answer(q, chunks, tenant)

        answer_id = f"a_{uuid.uuid4().hex[:12]}"
        answer = {
            "id": answer_id,
            "question_id": q["id"],
            "questionnaire_id": q_id,
            "tenant_id": tenant_id,
            "status": "draft",
            "generated_by": "ai",
            "created_at": datetime.utcnow().isoformat(),
            **answer_data,
        }

        if q_id not in answers_store:
            answers_store[q_id] = {}
        answers_store[q_id][q["id"]] = answer
        generated.append(answer)

    questionnaires[q_id]["progress"]["answered"] = len(answers_store.get(q_id, {}))
    questionnaires[q_id]["status"] = "in_review"

    return {"answers": generated, "total": len(generated)}


@app.post("/api/questionnaires/{q_id}/validate")
async def validate_answers(
    q_id: str,
    req: ValidateRequest,
    tenant_id: str = Depends(verify_tenant),
):
    """Validate generated answers."""
    if q_id not in questionnaires:
        raise HTTPException(status_code=404, detail="Questionnaire not found")
    if questionnaires[q_id]["tenant_id"] != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")

    qs = {q["id"]: q for q in questions_store.get(q_id, [])}
    ans = answers_store.get(q_id, {})

    if req.answer_ids:
        ans = {k: v for k, v in ans.items() if k in req.answer_ids}

    results = []
    for q_id_key, answer in ans.items():
        question = qs.get(q_id_key, {})
        chunks = await RetrievalAgent.search(question.get("text", ""), tenant_id, top_k=3)
        validation = await ValidatorAgent.validate_answer(question, answer, chunks)
        answer["status"] = "validated" if validation["status"] == "pass" else "needs_review"
        results.append({"answer_id": answer["id"], **validation})

    return {"validations": results, "total": len(results)}


@app.post("/api/questionnaires/{q_id}/approve/{question_id}")
async def approve_answer(
    q_id: str,
    question_id: str,
    tenant_id: str = Depends(verify_tenant),
):
    """Approve an answer (human-in-the-loop)."""
    if q_id not in questionnaires:
        raise HTTPException(status_code=404, detail="Questionnaire not found")
    if questionnaires[q_id]["tenant_id"] != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")

    ans = answers_store.get(q_id, {}).get(question_id)
    if not ans:
        raise HTTPException(status_code=404, detail="Answer not found")

    ans["status"] = "approved"
    ans["approved_at"] = datetime.utcnow().isoformat()
    return {"status": "approved", "answer_id": ans["id"]}


@app.post("/api/questionnaires/{q_id}/reject/{question_id}")
async def reject_answer(
    q_id: str,
    question_id: str,
    reason: str = "",
    tenant_id: str = Depends(verify_tenant),
):
    """Reject an answer."""
    if q_id not in questionnaires:
        raise HTTPException(status_code=404, detail="Questionnaire not found")
    if questionnaires[q_id]["tenant_id"] != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")

    ans = answers_store.get(q_id, {}).get(question_id)
    if not ans:
        raise HTTPException(status_code=404, detail="Answer not found")

    ans["status"] = "rejected"
    ans["rejection_reason"] = reason
    ans["rejected_at"] = datetime.utcnow().isoformat()
    return {"status": "rejected", "answer_id": ans["id"]}


@app.get("/api/questionnaires/{q_id}/export")
async def export_questionnaire(
    q_id: str,
    format: str = Query("excel", regex="^(excel|csv)$"),
    tenant_id: str = Depends(verify_tenant),
):
    """Export completed questionnaire."""
    if q_id not in questionnaires:
        raise HTTPException(status_code=404, detail="Questionnaire not found")
    if questionnaires[q_id]["tenant_id"] != tenant_id:
        raise HTTPException(status_code=403, detail="Access denied")

    qs = questions_store.get(q_id, [])
    ans = answers_store.get(q_id, {})

    if format == "excel":
        data = FormatterAgent.format_excel(questionnaires[q_id], qs, ans)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"questionnaire_{q_id}.xlsx"
    else:
        data = FormatterAgent.format_csv(questionnaires[q_id], qs, ans)
        media_type = "text/csv"
        filename = f"questionnaire_{q_id}.csv"

    from fastapi.responses import Response
    return Response(
        content=data,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/questionnaires")
async def list_questionnaires(tenant_id: str = Depends(verify_tenant)):
    """List all questionnaires for a tenant."""
    qs = [q for q in questionnaires.values() if q["tenant_id"] == tenant_id]
    return {"questionnaires": qs, "total": len(qs)}


# ============================================
# Knowledge Base Management
# ============================================

@app.post("/api/kb/ingest")
async def ingest_document(
    req: KBIngestRequest,
    tenant_id: str = Depends(verify_tenant),
):
    """Ingest a document into the knowledge base."""
    doc_id = f"doc_{uuid.uuid4().hex[:12]}"
    now = datetime.utcnow().isoformat()

    # Chunk the content
    chunks = []
    chunk_size = 1000
    overlap = 200
    text = req.content
    i = 0
    chunk_index = 0
    while i < len(text):
        end = min(i + chunk_size, len(text))
        chunk_text = text[i:end]

        # Get embedding
        embedding = await RetrievalAgent._get_embedding(chunk_text)

        chunk_id = f"chunk_{uuid.uuid4().hex[:12]}"
        chunk = {
            "id": chunk_id,
            "doc_id": doc_id,
            "tenant_id": tenant_id,
            "text": chunk_text,
            "embedding": embedding,
            "metadata": {"chunk_index": chunk_index, "start_char": i, "end_char": end},
        }
        chunks.append(chunk)

        # Store in Qdrant if embedding available
        if embedding:
            try:
                async with httpx.AsyncClient(timeout=10.0) as client:
                    await client.put(
                        f"{QDRANT_URL}/collections/kb-chunks/points",
                        json={
                            "points": [{
                                "id": chunk_id,
                                "vector": embedding,
                                "payload": {
                                    "text": chunk_text,
                                    "tenant_id": tenant_id,
                                    "doc_id": doc_id,
                                    "doc_name": req.name,
                                    "chunk_index": chunk_index,
                                },
                            }],
                        },
                    )
            except Exception as e:
                logger.error(f"Qdrant store error: {e}")

        i += chunk_size - overlap
        chunk_index += 1

    kb_docs[doc_id] = {
        "id": doc_id,
        "tenant_id": tenant_id,
        "name": req.name,
        "type": req.type,
        "description": req.description,
        "tags": req.tags,
        "chunk_count": len(chunks),
        "processing_status": "completed",
        "created_at": now,
    }

    if tenant_id not in kb_chunks:
        kb_chunks[tenant_id] = []
    kb_chunks[tenant_id].extend(chunks)

    return {
        "doc_id": doc_id,
        "name": req.name,
        "chunks_created": len(chunks),
        "status": "completed",
    }


@app.get("/api/kb/search")
async def search_kb(
    q: str,
    top_k: int = Query(5, ge=1, le=20),
    tenant_id: str = Depends(verify_tenant),
):
    """Search the knowledge base."""
    results = await RetrievalAgent.search(q, tenant_id, top_k)
    return {"results": results, "total": len(results)}


@app.get("/api/kb/documents")
async def list_kb_docs(tenant_id: str = Depends(verify_tenant)):
    """List knowledge base documents."""
    docs = [d for d in kb_docs.values() if d["tenant_id"] == tenant_id]
    return {"documents": docs, "total": len(docs)}


# ============================================
# Chat / Ask
# ============================================

@app.post("/api/chat")
async def chat(
    req: ChatRequest,
    tenant_id: str = Depends(verify_tenant),
):
    """Chat with the AI about security questionnaires."""
    chunks = await RetrievalAgent.search(req.message, tenant_id, top_k=3)
    context = "\n".join(c["text"][:300] for c in chunks)

    system = f"""You are a security compliance assistant. Help answer questions about security questionnaires and compliance.

Company: {tenants.get(tenant_id, {}).get('company_name', 'Unknown')}
Frameworks: {', '.join(tenants.get(tenant_id, {}).get('compliance_frameworks', []))}

Relevant context from knowledge base:
{context}

Be concise, accurate, and professional."""

    result = await AnswerAgent._call_llm(system, req.message, OMNI_URL)
    if not result:
        result = await AnswerAgent._call_llm(system, req.message, ROUTER9_URL)

    return {
        "response": result or "Unable to generate response",
        "citations": [{"doc_name": c.get("doc_name", ""), "excerpt": c["text"][:200]} for c in chunks],
    }


# ============================================
# Run
# ============================================

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", "8000"))
    uvicorn.run(app, host="0.0.0.0", port=port)
